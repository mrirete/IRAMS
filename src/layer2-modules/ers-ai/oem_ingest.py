"""
OEM Manual RAG Ingestion & Search Engine (Phase 5, Cap 6)
════════════════════════════════════════════════════════════

Ingests OEM equipment manuals (PDF/DOCX/TXT/XLSX) → chunks → embeddings → pgvector.
Provides RAG-powered search with source citations:
    "What's the torque spec for V-201 bonnet bolts?" → Answer + page reference.

Uses the Layer 3 RAG engine for chunking and embedding, with document metadata
(page numbers, sections, OEM name) preserved for citation accuracy.

Safety: Inherits the RAG engine's safety exclusion zone enforcement.
"""

import logging
import os
import io
from typing import Any, Dict, List, Optional

logger = logging.getLogger("ers.ai.oem_ingest")

# ── Document Extraction ──────────────────────────────────────


def _extract_pdf(file_bytes: bytes) -> List[Dict[str, Any]]:
    """Extract text from PDF, returning per-page chunks with page numbers."""
    pages = []
    try:
        import pdfplumber
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            for i, page in enumerate(pdf.pages, 1):
                text = page.extract_text()
                if text and text.strip():
                    pages.append({"text": text.strip(), "page": i})
    except ImportError:
        # Fallback to PyPDF2
        try:
            from PyPDF2 import PdfReader
            reader = PdfReader(io.BytesIO(file_bytes))
            for i, page in enumerate(reader.pages, 1):
                text = page.extract_text()
                if text and text.strip():
                    pages.append({"text": text.strip(), "page": i})
        except ImportError:
            logger.error("Neither pdfplumber nor PyPDF2 installed — cannot process PDFs")
    return pages


def _extract_docx(file_bytes: bytes) -> List[Dict[str, Any]]:
    """Extract text from DOCX file."""
    try:
        from docx import Document
        doc = Document(io.BytesIO(file_bytes))
        full_text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
        if full_text:
            return [{"text": full_text, "page": None}]
    except ImportError:
        logger.error("python-docx not installed — cannot process DOCX")
    return []


def _extract_txt(file_bytes: bytes) -> List[Dict[str, Any]]:
    """Extract text from plain text file."""
    text = file_bytes.decode("utf-8", errors="replace").strip()
    return [{"text": text, "page": None}] if text else []


def _extract_xlsx(file_bytes: bytes) -> List[Dict[str, Any]]:
    """Extract text from Excel file (concatenates all sheets)."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
        pages = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_text = []
            for row in ws.iter_rows(values_only=True):
                row_str = " | ".join(str(cell) for cell in row if cell is not None)
                if row_str.strip():
                    rows_text.append(row_str)
            if rows_text:
                pages.append({"text": f"Sheet: {sheet_name}\n" + "\n".join(rows_text), "page": None})
        return pages
    except ImportError:
        logger.error("openpyxl not installed — cannot process XLSX")
    return []


# ── File Type Router ─────────────────────────────────────────

_EXTRACTORS = {
    ".pdf": _extract_pdf,
    ".docx": _extract_docx,
    ".txt": _extract_txt,
    ".xlsx": _extract_xlsx,
    ".xls": _extract_xlsx,
}


# ── OEM RAG Engine ───────────────────────────────────────────

class OEMRAGEngine:
    """
    OEM Manual RAG Engine — ingests equipment manuals and provides
    retrieval-augmented search with source citations.
    """

    def __init__(self):
        # Import the base RAG engine from Layer 3
        try:
            from layer3_agents.engines.rag import RAGEngine
            self._rag = RAGEngine(chunk_size=512)
        except ImportError:
            # Fallback: inline lightweight chunker
            self._rag = None
            self._chunks: list = []

        self._documents: List[Dict[str, Any]] = []
        logger.info("OEM RAG Engine initialized")

    def ingest_file(
        self,
        file_bytes: bytes,
        filename: str,
        asset_tag: Optional[str] = None,
        equipment_class: Optional[str] = None,
        document_type: str = "oem_manual",
    ) -> int:
        """
        Ingest a file into the RAG vector store.
        Returns the number of chunks created.
        """
        ext = os.path.splitext(filename)[1].lower()
        extractor = _EXTRACTORS.get(ext)
        if not extractor:
            raise ValueError(f"Unsupported file type: {ext}. Supported: {', '.join(_EXTRACTORS.keys())}")

        pages = extractor(file_bytes)
        if not pages:
            raise ValueError(f"No text content extracted from '{filename}'")

        total_chunks = 0
        for page_data in pages:
            source = filename
            if page_data.get("page"):
                source = f"{filename} (p.{page_data['page']})"

            if self._rag:
                count = self._rag.ingest_document(page_data["text"], source)
            else:
                # Lightweight fallback
                self._chunks.append({
                    "text": page_data["text"],
                    "source": source,
                    "page": page_data.get("page"),
                    "asset_tag": asset_tag,
                    "equipment_class": equipment_class,
                })
                count = 1

            total_chunks += count

        # Track document metadata
        self._documents.append({
            "filename": filename,
            "document_type": document_type,
            "asset_tag": asset_tag,
            "equipment_class": equipment_class,
            "chunk_count": total_chunks,
            "page_count": len(pages),
        })

        logger.info(
            "Ingested '%s': %d pages, %d chunks (asset=%s, class=%s)",
            filename, len(pages), total_chunks, asset_tag, equipment_class,
        )
        return total_chunks

    def query_oem(
        self,
        query: str,
        asset_tag: Optional[str] = None,
        equipment_class: Optional[str] = None,
        top_k: int = 5,
    ) -> Dict[str, Any]:
        """
        RAG-powered OEM manual search.
        Returns answer with source citations.
        """
        if self._rag:
            rag_result = self._rag.query(query, top_k=top_k)

            if rag_result.safety_blocked:
                return {
                    "answer": "",
                    "sources": [],
                    "safety_blocked": True,
                    "safety_reason": rag_result.safety_reason,
                }

            # Generate answer using Gemini if available
            answer = rag_result.answer
            try:
                # Use Gemini to synthesize a proper answer from retrieved chunks
                from .service import _get_client, GEMINI_MODEL, RELANTERN_SYSTEM_INSTRUCTION
                from google.genai import types

                client = _get_client()
                context = "\n\n".join(
                    f"[Source: {s.get('document', 'Unknown')}]\n{s.get('excerpt', '')}"
                    for s in rag_result.sources
                )

                synthesis_prompt = f"""Based on the following OEM manual excerpts, answer the user's question accurately.
Include specific values, specifications, and page references where available.
If the information is not found in the excerpts, say so clearly.

User Question: {query}

Relevant Manual Excerpts:
{context}

Provide a clear, technical answer with source citations."""

                response = client.models.generate_content(
                    model=GEMINI_MODEL,
                    contents=synthesis_prompt,
                    config=types.GenerateContentConfig(
                        system_instruction=RELANTERN_SYSTEM_INSTRUCTION +
                        "\n\nYou are answering questions from OEM equipment manuals. Be precise and cite sources.",
                        temperature=0.2,
                    ),
                )
                answer = response.text or answer
            except Exception as e:
                logger.warning("Gemini synthesis failed, using raw RAG answer: %s", e)

            return {
                "answer": answer,
                "sources": rag_result.sources,
                "safety_blocked": False,
            }
        else:
            # Fallback: simple keyword search
            query_terms = set(query.lower().split())
            scored = []
            for chunk in self._chunks:
                # Filter by asset/class if specified
                if asset_tag and chunk.get("asset_tag") and chunk["asset_tag"] != asset_tag:
                    continue
                if equipment_class and chunk.get("equipment_class") and chunk["equipment_class"] != equipment_class:
                    continue

                chunk_terms = set(chunk["text"].lower().split())
                overlap = len(query_terms & chunk_terms)
                if overlap > 0:
                    score = overlap / max(len(query_terms), 1)
                    scored.append((chunk, score))

            scored.sort(key=lambda x: x[1], reverse=True)
            top = scored[:top_k]

            sources = [
                {
                    "document": c["source"],
                    "page": c.get("page"),
                    "excerpt": c["text"][:200] + "..." if len(c["text"]) > 200 else c["text"],
                    "score": round(s, 3),
                }
                for c, s in top
            ]

            if top:
                answer = f"Based on {len(top)} sources: {top[0][0]['text'][:300]}..."
            else:
                answer = "No relevant documents found for your query."

            return {
                "answer": answer,
                "sources": sources,
                "safety_blocked": False,
            }

    def list_documents(self) -> List[Dict[str, Any]]:
        """List all ingested documents."""
        return self._documents


# ── Singleton Instance ───────────────────────────────────────
oem_rag_engine = OEMRAGEngine()
